from openpyxl import load_workbook
from address import address_remove
from bairros import cepBairros
from ruas import RUAS


def path_project():
  from pathlib import Path
  
  return Path(__file__).parent.absolute()

def load_file(filename: str):
  # file = path_project(filename)

  workbook = load_workbook(filename)
  worksheet = workbook.active

  return [workbook, worksheet]

def copy_sheet(ws) -> list:
  maxColumns = ws.max_column
  maxRows = ws.max_row
  copiedSheet = []

  for row in ws.iter_rows(max_row=maxRows,max_col=maxColumns, values_only=True):
    accumulated_items = []
    for item in row:
      accumulated_items.append(item)

    copiedSheet.append(accumulated_items)

  return copiedSheet

def index_column_address(data, address='Destination Address'):
  for index, item in enumerate(data[0]):
    if item == address:
      return index

def split_address(data, columnAddress, columnDescription):
  for indexRow, row in enumerate(data[1::], 1):
    addressData = data[indexRow][columnAddress]
    [address, description] = split_destination(addressData)

    data[indexRow][columnAddress] = address
    data[indexRow][columnDescription] = description

def split_destination(destination: str):
  from re import sub

  # removendo excesso de espaço
  destination = sub(r"\s\s+", " ", destination)

  [street, number, *description] = destination.strip().split(', ')

  if len(description) == 0:
    description = None
  else:
    description = description[0]

  address = f'{street}, {number}'

  return [address, description]

def insert_column(data, column, description, valueInitial=''):
  """
    Insere uma coluna nova na matriz e coloca '' no campo
  """

  for indexRow, row in enumerate(data, 0):
    data[indexRow].insert(column, valueInitial)

  data[0][column] = description


def remove_columns(ws, data, listColumns: list):
  for row in listColumns:
    idx = data[0].index(row)

    ws.delete_cols(idx)

def organize_worksheet(data, columnDescription, columnAddress):
  """
    Função responsável por:
      1. Seperar o endereço em 2 partes:
        1: Rua e numero
        2: Referência
        
      2. Retirar o 'R ' no inicio do endereço substituindo por 'Rua '
      
  """

  # insere uma nova coluna
  insert_column(data, columnDescription, 'Address Line 2')

  # insere os endereços na nova coluna
  split_address(data, columnAddress, columnDescription)
  data[0][columnAddress] = 'Address Line 1'


def check_address(data: list, col: int):
  """Remove os endereços incorretos e insere os certos"""
  import re
  REGEX = r'^(r\. |r |rua )'

  for idxRow, row in enumerate(data[1::],1):
    adrs:str = row[col].lower()
    
    pattern = re.compile(REGEX, re.IGNORECASE)

    [address, number] = adrs.split(', ')
    
    if pattern.match(adrs):
      address = pattern.sub('Rua ', address)
      data[idxRow][col] = f'{address.title()}, {number}'

  return data


def verifyStreetNameByZipCode(address: str, listInfoAddress: list):
    '''Corrige o nome do endereço de acordo com o CEP informado'''
    index_cep =  listInfoAddress[0].index('Zipcode/Postal code')
    
    cep = address[index_cep].replace('-', '')
    newAddress = cepBairros.get(cep)
    
    if newAddress is None:
      idx_logradouro =  listInfoAddress[0].index('Address Line 1')
      return address[idx_logradouro]
    
    return newAddress

# IGNORAR
def verifyStreetNameByZipCode2(data: list, col: list):
  index_cep =  data[0].index('Zipcode/Postal code')
  index_logradouro =  data[0].index('Address Line 1')

  for idx_linha, linha in enumerate(data[1::], 1):
    cep = data[idx_linha][index_cep].replace('-', '')
    numero = data[idx_linha][index_logradouro].split(', ')[1]

    logradouro = cepBairros.get(cep)

    if logradouro:
      data[idx_linha][index_logradouro] = f'{logradouro}, {numero}'



def buscaCEP(data: list):
  import requests

  visited_postal_codes = []

  index_cep =  data[0].index('Zipcode/Postal code')
  index_logradouro =  data[0].index('Address Line 1')

  for idx_linha, linha in enumerate(data[1::], 1):
    cep = data[idx_linha][index_cep]

    if cep in visited_postal_codes:
      continue

    URL = f'https://viacep.com.br/ws/{cep}/json/'
    results: dict = requests.get(URL).json()

    print(URL)
    visited_postal_codes.append(cep)

    logradouro = results.get('logradouro', '')
    numero = data[idx_linha][index_logradouro].split(', ')[1]


    if logradouro:
      data[idx_linha][index_logradouro] = f'{logradouro}, {numero}'



'''
ED 1941
ED 1945

ED 1877
ED 1879
ED 1881
'''


def groupAddress(data: list):
  pass

  

"""
def fetchAddress(data: list, col: int):
  import requests
  from os import getenv
  from dotenv import load_dotenv
  from urllib.parse import quote
  
  load_dotenv()
  GOOGLE = getenv('GOOGLE')
  
  for idxRow, row in enumerate(data[1::],1):
    street = quote(row[col])
    print(row[col])
    
    cityIndex = data[0].index('City')
    city = row[cityIndex]

    try:
      URL = f'https://maps.googleapis.com/maps/api/geocode/json?address={street}&components=locality:{city}&key={GOOGLE}'
      results = requests.get(URL).json()['results']

      if results:
        address_components = results[0].get('address_components', True)
        isPolitical = 'politicial' in address_components[0].get('types', True)

        if isPolitical:
          continue
    
      else:
        autocomplete_url = f"https://maps.googleapis.com/maps/api/place/autocomplete/json?input={street}&location=-18.9439,-46.9929&radius=5000&key={GOOGLE}"
        autocomplete_result = requests.get(autocomplete_url).json()

        if autocomplete_result:
          place_id = autocomplete_result['predictions'][0]['place_id']
          details_url = f"https://maps.googleapis.com/maps/api/place/details/json?place_id={place_id}&key={GOOGLE}"

        details_result = requests.get(details_url).json()['result']

        location = details_result['geometry']['location']
        # lat
        data[idxRow][-2] = location['lat']

        # lng
        data[idxRow][-1] = location['lng']
    
    except Exception as error:
      print(error)
      
    
    print(idxRow) # LOGGING

  return data
"""


def edit_length(data: list, col):
  """Remove o '1' na coluna de quantidade e o deixa sem nada"""
  for line, row in enumerate(data[1::], 1):
    if row[col] == 1:
      data[line][col] = ''


def correct_address(data: list, idxColAddress: str):
  from rapidfuzz import process, fuzz

  for idxRow, line in enumerate(data[1::], 1):
    address, number = line[idxColAddress].split(', ')

    newAddress, score, _ = process.extractOne(address, RUAS, scorer=fuzz.token_set_ratio)

    if score < 85.0:
      # novo endereço
      newAddress = verifyStreetNameByZipCode(line, data)

    line[idxColAddress] = f'{newAddress}, {number}'

def search_duplicated(data: list, address: str, columnAddress: str):
  """
    Retorna o index de um endereço em uma lista de outros endereços
  """
  for indexRow, row in enumerate(data[1::], 1):
    if(address in row[columnAddress]):
      return indexRow


def remove_duplicate(data, columnAddress, columnCount):
  """
    Aumenta a quantidade nos endereço repetido e remove os outros
  """

  newData = []
  listAddress = []
  newData.insert(0, data[0])

  idxRow = 1
  for indexRow, row in enumerate(data[1::], 1):
    address = row[columnAddress].title()
    row[columnAddress] = address # deixando maiusculo o nome da rua
    countAddressOnList = listAddress.count(address)

    # Não existe na lista
    if(countAddressOnList == 0):
      newData.append(data[indexRow])
      listAddress.append(address)

      q = newData[idxRow][columnCount]
      idxRow += 1

      continue

    # Existe na lista de paradas(add +1)
    indxSD = search_duplicated(newData, address, columnAddress)
    q = newData[indxSD][columnCount]
    newData[indxSD][columnCount] = 1 + q

  return newData

def addAddress(data: list):
  # POR DO SOL
  
  por_do_sol = ['-', 'Pôr Do Sol', '', '', 'Congonha', 'Patrocínio', '38740-000', '-18.92639', '-47.00644']
  data.append(por_do_sol)

  return data

def remove_apartment(data, columnAddress, columnLength):
  """
    Remove alguns endereços especificados no arquivos 'address'
  """
  
  import re

  indexRow = 1
  countAddress = 0

  for row in data[1::]:
    addressLowerCase: str = row[columnAddress].lower()
    address = re.sub(r'^(rua |r )', '', addressLowerCase.split(', ')[0])

    addressRemove = address_remove.get(address)

    if(addressRemove != None):
      del data[indexRow]
      countAddress += row[columnLength]

    else:
      indexRow += 1

  if(countAddress):
    data = addAddress(data)
    data[-1][columnLength] = countAddress
    
  return data


def save_workbook(data, file):
  """salva as informações no arquivo"""

  from openpyxl import Workbook

  workbook = Workbook()
  worksheet = workbook.active

  for indexRow, row in enumerate(data):
    for indexColumn, column in enumerate(row):
      value = column

      worksheet.cell(row=indexRow+1, column=indexColumn+1, value=value)

  workbook.save(file)


def formatDateFile():
  '''
  Retorna o dia da semana com o sufixo
  '''

  from datetime import datetime
  from pytz import timezone
  import locale

  # Configura a localidade para português do Brasil
  locale.setlocale(locale.LC_TIME, 'pt_BR.utf8')
  tz = timezone('America/Sao_Paulo')

  day = datetime.now(tz).strftime("%A")

  if day in ['segunda', 'terça', 'quarta', 'quinta', 'sexta']:
    day += '-feira'

  return f'{day}.xlsx'


def permitionUser(idUser):
  with open('users.txt', 'r') as file:
    while True:
      user = file.readline().replace('\n', '')
      
      if user:
        if user == idUser:
          return True
        
        continue
        
      return False
